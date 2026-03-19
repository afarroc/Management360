# analyst/services/excel_analyzer.py
"""
Detección automática de regiones de datos en hojas Excel.
Incluye detección de grupos fusionables (mismo ancho de columnas,
separados solo por filas vacías → probable dataset partido).
"""

import logging
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)

MAX_GAP_TO_MERGE = 5  # filas vacías de separación máxima para considerar fusión


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _col_letter(n):
    return get_column_letter(n)

def _range_str(r1, c1, r2, c2):
    return f"{_col_letter(c1)}{r1}:{_col_letter(c2)}{r2}"

def _cell_has_value(cell):
    v = cell.value
    if v is None: return False
    if isinstance(v, str) and v.strip() == '': return False
    return True

def _build_matrix(ws):
    """
    Construye la matriz booleana de celdas con dato.

    Usa ws.iter_rows() en lugar de ws.cell(row, col) individual.
    En modo read_only, iter_rows() recorre el XML UNA sola vez (O(n)),
    mientras que ws.cell() lo hace por cada celda (O(n²) efectivo).
    Para la hoja 'Datos' (301×9): iter_rows = 0.016 s vs ws.cell = 23.8 s.
    """
    matrix = []
    for row in ws.iter_rows():
        matrix.append([
            v.value is not None and not (isinstance(v.value, str) and v.value.strip() == '')
            for v in row
        ])
    return matrix

def _density(matrix, r1, c1, r2, c2):
    total = (r2 - r1 + 1) * (c2 - c1 + 1)
    if total == 0: return 0.0
    filled = sum(1 for r in range(r1-1, r2) for c in range(c1-1, c2) if matrix[r][c])
    return filled / total

def _make_region(r1, c1, r2, c2, matrix, idx):
    rows = r2 - r1 + 1
    cols = c2 - c1 + 1
    dens = _density(matrix, r1, c1, r2, c2)
    score = dens * rows * cols
    confidence = round(dens * 0.65 + min(1.0, (rows * cols) / 25) * 0.35, 3)
    return {
        'range_str':   _range_str(r1, c1, r2, c2),
        'start_row': r1, 'start_col': c1,
        'end_row':   r2, 'end_col':   c2,
        'rows': rows, 'cols': cols,
        'density':    round(dens, 3),
        'score':      round(score, 3),
        'confidence': confidence,
        'label':      f"Región {idx}",
        'description': f"{_col_letter(c1)}{r1}:{_col_letter(c2)}{r2}  —  {rows} filas × {cols} cols  —  densidad {dens:.0%}",
    }


# ──────────────────────────────────────────────────────────────────────────────
# Detección de regiones (con soporte de tablas esparsas con cabecera)
# ──────────────────────────────────────────────────────────────────────────────

def _best_header_fill(matrix, r1, c1, r2, c2, scan_rows=3) -> tuple:
    """
    Escanea las primeras `scan_rows` filas del bloque y devuelve:
      (max_fill, header_row_1indexed)

    Esto resuelve el caso de celdas fusionadas encima de la cabecera real:
      Fila 6: celdas fusionadas C6:G6 / H6:L6  → openpyxl solo llena C6 y H6 → fill ~18%
      Fila 7: B7:L7 con cabeceras reales         → fill 100%
    El scan devuelve (1.0, 7) y la fila 7 se usa como el inicio lógico de la tabla.
    """
    cols = c2 - c1 + 1
    if cols == 0:
        return 0.0, r1
    best_fill, best_row = 0.0, r1
    limit = min(r1 + scan_rows - 1, r2)        # no salir del bloque
    for r in range(r1, limit + 1):
        filled = sum(1 for c in range(c1-1, c2) if matrix[r-1][c])  # c1-1..c2-1 (0-indexed)
        fill = filled / cols
        if fill > best_fill:
            best_fill, best_row = fill, r
    return best_fill, best_row


def _classify_region(r1, c1, r2, c2, matrix, idx) -> dict:
    """
    Construye el dict de región extendido con campos de tabla esparsa.

    Una región es 'sparse' cuando:
      - Densidad global < 35%
      - Alguna de las primeras 3 filas tiene ≥ 50% de fill (cabecera real)
      - Hay al menos 1 dato en el cuerpo (filas posteriores a la cabecera)

    Maneja el caso de filas de título con celdas fusionadas encima de la cabecera:
      Fila 6: celdas fusionadas (fill bajo) → se ignora
      Fila 7: cabecera real con todos los campos (fill 100%) → se detecta
    """
    base = _make_region(r1, c1, r2, c2, matrix, idx)
    cols = c2 - c1 + 1

    # Buscar la mejor fila de cabecera en las primeras 3 filas del bloque
    header_fill, header_row = _best_header_fill(matrix, r1, c1, r2, c2, scan_rows=3)

    # Cuerpo: filas después de la cabecera
    body_start = header_row + 1
    body_rows  = r2 - header_row
    if body_rows > 0:
        body_filled  = sum(1 for r in range(body_start - 1, r2) for c in range(c1-1, c2) if matrix[r][c])
        body_density = body_filled / (body_rows * cols)
    else:
        body_density = 0.0

    is_sparse = base['density'] < 0.35 and header_fill >= 0.5

    base['sparse']       = is_sparse
    base['header_fill']  = round(header_fill, 3)
    base['header_row']   = header_row          # fila 1-indexed de la cabecera real
    base['body_density'] = round(body_density, 3)

    if is_sparse:
        filled_cols = sum(
            1 for c in range(c1-1, c2)
            if any(matrix[r][c] for r in range(body_start - 1, r2))
        )
        base['description'] = (
            f"{base['range_str']}  —  {base['rows']} filas × {cols} cols  "
            f"—  cabecera fila {header_row} ({header_fill:.0%} llena), "
            f"{filled_cols} col{'s' if filled_cols!=1 else ''} con datos"
        )
    return base


def _find_all_regions(matrix, min_density=0.4, min_cells=2):
    """
    Detecta TODAS las regiones rectangulares independientes en la hoja.

    Criterios de aceptación:
      A) Densidad global ≥ min_density  (tabla densa normal)
      B) Alguna de las primeras 3 filas tiene ≥ 50% fill (cabecera) +
         al menos 1 dato en el cuerpo posterior
         → tabla esparsa (ej. B6:L81 donde fila 6 tiene merged cells de título
           y fila 7 tiene cabeceras B7:L7, cuerpo solo col B)
    """
    if not matrix: return []
    max_row = len(matrix)
    max_col = len(matrix[0]) if matrix else 0

    def row_groups():
        groups, in_g, g_start = [], False, 0
        for r in range(max_row):
            has = any(matrix[r])
            if has and not in_g:    in_g, g_start = True, r
            elif not has and in_g:  in_g = False; groups.append((g_start, r - 1))
        if in_g: groups.append((g_start, max_row - 1))
        return groups

    def col_groups(rb_start, rb_end):
        col_active = [any(matrix[r][c] for r in range(rb_start, rb_end + 1)) for c in range(max_col)]
        segs, in_g, g_start = [], False, 0
        for c in range(max_col):
            if col_active[c] and not in_g:    in_g, g_start = True, c
            elif not col_active[c] and in_g:  in_g = False; segs.append((g_start, c - 1))
        if in_g: segs.append((g_start, max_col - 1))
        return segs

    def accept(r1, c1, r2, c2):
        area = (r2 - r1 + 1) * (c2 - c1 + 1)
        if area < min_cells:
            return False
        dens = _density(matrix, r1, c1, r2, c2)
        if dens >= min_density:
            return True                                  # A) tabla densa
        # B) tabla esparsa: alguna cabecera densa en las primeras 3 filas
        hfill, hrow = _best_header_fill(matrix, r1, c1, r2, c2, scan_rows=3)
        if hfill >= 0.5 and r2 > hrow:                  # hay cuerpo después de la cabecera
            return any(matrix[r][c] for r in range(hrow, r2) for c in range(c1-1, c2))
        return False

    regions = []
    for (rb_start, rb_end) in row_groups():
        for (cb_start, cb_end) in col_groups(rb_start, rb_end):
            r1, c1 = rb_start + 1, cb_start + 1
            r2, c2 = rb_end + 1,   cb_end + 1
            if accept(r1, c1, r2, c2):
                regions.append(_classify_region(r1, c1, r2, c2, matrix, len(regions) + 1))

    regions.sort(key=lambda x: x['score'], reverse=True)
    for i, reg in enumerate(regions):
        reg['label'] = f"Región {i + 1}"
    return regions


# ──────────────────────────────────────────────────────────────────────────────
# Detección de grupos fusionables
# ──────────────────────────────────────────────────────────────────────────────

def _find_merge_groups(regions, matrix):
    """
    Agrupa regiones que son candidatas a ser un solo dataset partido por filas vacías.

    Criterios para que dos regiones sean fusionables:
      1. Mismo rango de columnas (start_col y end_col iguales) → mismo ancho.
      2. Separación vertical ≤ MAX_GAP_TO_MERGE filas vacías.
      3. Ordenadas por posición vertical (start_row).

    Retorna una lista de grupos. Cada grupo es:
    {
        'region_indices': [i, j, ...],   # índices en la lista `regions`
        'merged_range':   str,           # rango fusionado "B3:D12"
        'merged_rows':    int,
        'gap_rows':       int,           # filas vacías entre primera y última región
        'reason':         str,           # texto explicativo para el usuario
        'confidence':     float,
    }
    Solo se incluyen grupos de ≥ 2 regiones.
    """
    if len(regions) < 2:
        return []

    # Ordenar por posición vertical para comparar adyacentes
    sorted_regs = sorted(enumerate(regions), key=lambda x: x[1]['start_row'])

    # Agrupar por columnas idénticas
    # Clave de columna: (start_col, end_col)
    from collections import defaultdict
    by_cols = defaultdict(list)
    for orig_idx, reg in sorted_regs:
        key = (reg['start_col'], reg['end_col'])
        by_cols[key].append((orig_idx, reg))

    merge_groups = []

    for (c1, c2), col_group in by_cols.items():
        if len(col_group) < 2:
            continue

        # Dentro de este grupo de columnas, agrupar por proximidad vertical
        # usando ventana deslizante
        current_chain = [col_group[0]]
        chains = []

        for i in range(1, len(col_group)):
            prev_orig_idx, prev_reg = current_chain[-1]
            curr_orig_idx, curr_reg = col_group[i]
            gap = curr_reg['start_row'] - prev_reg['end_row'] - 1  # filas vacías entre ellas

            if 0 <= gap <= MAX_GAP_TO_MERGE:
                current_chain.append(col_group[i])
            else:
                if len(current_chain) >= 2:
                    chains.append(current_chain)
                current_chain = [col_group[i]]

        if len(current_chain) >= 2:
            chains.append(current_chain)

        for chain in chains:
            indices   = [x[0] for x in chain]
            regs_list = [x[1] for x in chain]

            r1_merged = min(r['start_row'] for r in regs_list)
            r2_merged = max(r['end_row']   for r in regs_list)
            total_gap = sum(
                regs_list[i+1]['start_row'] - regs_list[i]['end_row'] - 1
                for i in range(len(regs_list) - 1)
            )
            total_data_rows = sum(r['rows'] for r in regs_list)
            merged_rows = r2_merged - r1_merged + 1

            # Densidad del rango fusionado (incluye las filas vacías del gap)
            merged_dens = _density(matrix, r1_merged, c1, r2_merged, c2)

            # Texto de razón para el usuario
            n = len(regs_list)
            gap_desc = f"{total_gap} fila{'s' if total_gap != 1 else ''} vacía{'s' if total_gap != 1 else ''}"
            reason = (
                f"{n} bloques con el mismo ancho de columnas ({_col_letter(c1)}→{_col_letter(c2)}), "
                f"separados por {gap_desc}. "
                f"Probablemente un solo dataset de {total_data_rows} filas de datos."
            )

            # Confianza de fusión: baja si el gap es grande, alta si es 1-2 filas
            gap_penalty = max(0, 1 - total_gap / (MAX_GAP_TO_MERGE * len(regs_list)))
            col_match   = 1.0  # siempre coincide porque es el criterio
            merge_conf  = round((gap_penalty * 0.6 + merged_dens * 0.4), 3)

            merge_groups.append({
                'region_indices': indices,
                'region_labels':  [r['label'] for r in regs_list],
                'merged_range':   _range_str(r1_merged, c1, r2_merged, c2),
                'start_row': r1_merged, 'start_col': c1,
                'end_row':   r2_merged, 'end_col':   c2,
                'merged_rows':    merged_rows,
                'merged_cols':    c2 - c1 + 1,
                'data_rows':      total_data_rows,
                'gap_rows':       total_gap,
                'density':        round(merged_dens, 3),
                'confidence':     merge_conf,
                'reason':         reason,
                'label':          f"Fusión de {' + '.join(r['label'] for r in regs_list)}",
                'description':    f"{_range_str(r1_merged, c1, r2_merged, c2)}  —  {total_data_rows} filas de datos + {total_gap} filas vacías",
            })

    return merge_groups


# ──────────────────────────────────────────────────────────────────────────────
# Clase principal
# ──────────────────────────────────────────────────────────────────────────────

class ExcelRangeAnalyzer:

    @classmethod
    def analyze_sheet(cls, file_obj, sheet_name: str) -> dict:
        """
        Retorna:
          regions       : todas las regiones independientes ordenadas por score
          merge_groups  : grupos de regiones fusionables (pueden estar vacío)
          recommended   : región de mayor score
          multi         : True si hay más de 1 región
        """
        try:
            file_obj.seek(0)
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

            if sheet_name not in wb.sheetnames:
                return {'error': f"Hoja '{sheet_name}' no encontrada", 'regions': [], 'merge_groups': []}

            ws = wb[sheet_name]
            max_row, max_col = ws.max_row or 0, ws.max_column or 0

            if max_row == 0 or max_col == 0:
                wb.close()
                return {'sheet': sheet_name, 'max_row': 0, 'max_col': 0,
                        'regions': [], 'merge_groups': [], 'recommended': None,
                        'multi': False, 'error': 'Hoja vacía'}

            matrix = _build_matrix(ws)
            wb.close()

            regions = _find_all_regions(matrix)

            if not regions:
                return {'sheet': sheet_name, 'max_row': max_row, 'max_col': max_col,
                        'regions': [], 'merge_groups': [], 'recommended': None,
                        'multi': False, 'error': 'No se detectaron regiones con datos'}

            merge_groups = _find_merge_groups(regions, matrix) if len(regions) > 1 else []

            return {
                'sheet':        sheet_name,
                'max_row':      max_row,
                'max_col':      max_col,
                'regions':      regions,
                'merge_groups': merge_groups,
                'recommended':  regions[0],
                'multi':        len(regions) > 1,
                'error':        None
            }

        except InvalidFileException:
            return {'error': 'Archivo Excel inválido o corrupto', 'regions': [], 'merge_groups': []}
        except Exception as e:
            logger.error(f"Error analizando hoja: {e}", exc_info=True)
            return {'error': str(e), 'regions': [], 'merge_groups': []}

    @classmethod
    def get_sheets(cls, file_obj) -> dict:
        try:
            file_obj.seek(0)
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
            return {'sheets': sheets, 'error': None}
        except Exception as e:
            logger.error(f"Error obteniendo hojas: {e}", exc_info=True)
            return {'sheets': [], 'error': str(e)}
